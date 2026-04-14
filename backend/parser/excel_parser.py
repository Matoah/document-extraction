from pathlib import Path
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from model.standard_specification_config import StandardSpecificationConfig


def parse_excel(excel_file: str | Path) -> list[StandardSpecificationConfig]:
    """解析excel文件"""
    excel_file_path = excel_file
    if isinstance(excel_file, str):
        excel_file_path = Path(excel_file).resolve()
    if not excel_file_path.exists():
        raise FileNotFoundError(f"excel文件 {excel_file} 不存在")
    if not excel_file_path.is_file():
        raise ValueError(f"excel文件 {excel_file} 不是文件")
    if not excel_file_path.suffix == ".xlsx":
        raise ValueError(f"excel文件 {excel_file} 不是xlsx文件")
    # 解析excel文件
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook.active
    row_count = sheet.max_row
    standard_items = []
    row = 3
    while row <= row_count:
        code_cell = sheet.cell(row=row, column=3)
        code = _get_merged_cell_value(sheet, code_cell)
        if code:
            plate = _get_merged_cell_value(sheet, sheet.cell(row=row, column=1))
            module = _get_merged_cell_value(sheet, sheet.cell(row=row, column=2))
            name = _get_merged_cell_value(sheet, sheet.cell(row=row, column=4))
            chief_org = _get_merged_cell_value(sheet, sheet.cell(row=row, column=5))
            chief_editor = _get_merged_cell_value(sheet, sheet.cell(row=row, column=6))
            contact_name = _get_merged_cell_value(sheet, sheet.cell(row=row, column=7))
            contact_phone = _get_merged_cell_value(sheet, sheet.cell(row=row, column=8))
            contact_email = _get_merged_cell_value(sheet, sheet.cell(row=row, column=9))
            contact_address = _get_merged_cell_value(sheet, sheet.cell(row=row, column=10))
            file_names = []
            merged_row_count = _get_merged_row_count(sheet, code_cell)
            for delta in range(0, merged_row_count):
                file_name = _get_merged_cell_value(sheet, sheet.cell(row=row+delta, column=11))
                file_names.append(file_name)
            standard_items.append(StandardSpecificationConfig(
                plate=plate,
                module=module,
                code=code,
                name=name,
                chief_org=chief_org,
                chief_editor=chief_editor,
                contact_name=contact_name,
                contact_phone=contact_phone,
                contact_email=contact_email,
                contact_address=contact_address,
                file_names=file_names,
            ))
            row += merged_row_count
        else:
            row += 1
    return standard_items


def _get_merged_row_count(sheet: Worksheet, cell: Cell) -> int:
    """获取合并单元格的行数"""
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            return max_row - min_row + 1
    return 1

def _get_merged_cell_value(ws: Worksheet, cell: Cell):
    """获取合并单元格的值"""
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 左上角单元格
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


if __name__ == "__main__":
    standard_items = parse_excel(str(Path(__file__).parent.parent / "现行公路工程行业标准.xlsx"))
    print(standard_items)

